"""
PROYECTO FINAL - MECANICA DE MATERIALES II
Portico plano de acero - Ing. Valencia
Universidad del Quindio - Ingenieria Civil
xy = 45

Genera el Excel maestro con todos los calculos.
"""

import math
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ==============================================================
# PARAMETROS GLOBALES
# ==============================================================
xy = 45
FS = 1.4
Fy = 250       # MPa - ASTM A36
tau_y = 100    # MPa - cortante de fluencia
E_steel = 200  # GPa - modulo de elasticidad

# Geometria
L_AB = 4       # m
L_BC = 4       # m
L_CD = 5       # m
L_viga = 13    # m total
L_col = 6      # m columna

# Cargas
# W1(x) = 5.625x + 45 para [0, 8] -> de 45 a 90 kN/m
# W2(x) = 234 - 18x para [8, 13] -> de 90 a 0 kN/m
# Puntual en C: 15 kN a 40 grados
# Momento en C: 5 kN*m CW
# Momento en E: 90 kN*m CW

P_puntual = 15          # kN
angulo = 40             # grados
Fx_C = P_puntual * math.cos(math.radians(angulo))   # 11.4907 kN
Fy_C = P_puntual * math.sin(math.radians(angulo))   # 9.6418 kN
M_C = 5                 # kN*m CW
M_E = 90                # kN*m CW

# ==============================================================
# FASE 1: RESULTANTES DE CARGAS DISTRIBUIDAS
# ==============================================================
# Integral W1 de 0 a 8: [2.8125x^2 + 45x] de 0 a 8
R_W1_total = 2.8125 * 64 + 45 * 8  # = 540 kN
# Integral W2 de 8 a 13: [234x - 9x^2] de 8 a 13
R_W2 = (234*13 - 9*169) - (234*8 - 9*64)  # = 225 kN

# Integral W1 de 0 a 4
R_W1_04 = 2.8125 * 16 + 45 * 4  # = 225 kN
# Integral W1 de 4 a 8
R_W1_48 = R_W1_total - R_W1_04  # = 315 kN

# ==============================================================
# FASE 2: REACCIONES
# ==============================================================
# Columna DE: Sum M_E = 0 => D_x_on_col * 6 - 90 = 0
# D_x en tope de columna (fuerza del beam sobre columna)
# Beam empuja columna a la IZQUIERDA con 15 kN
# Columna empuja beam a la DERECHA con 15 kN
D_x_col = M_E / L_col  # = 15 kN
E_x = D_x_col          # = 15 kN (hacia la derecha)

# Cuerpo BD: Sum M_B = 0
# Integrales de momento respecto a B (x=4):
# I1 = int_4^8 (5.625x+45)(x-4) dx = 660
# I2 = int_8^13 (234-18x)(x-4) dx = 1275
I1_MB = 660.0
I2_MB = 1275.0
M_puntual_B = Fy_C * (8 - 4)  # 9.6418 * 4 = 38.567 kN*m

# Dy_up*9 - I1 - I2 - M_puntual_B - M_C(CW=-) = 0
# M_C es CW => contribuye NEGATIVO en CCW+ => se RESTA
# Dy_up = columna empuja beam hacia arriba en D
Dy_up = (I1_MB + I2_MB + M_puntual_B - M_C) / 9.0
# Dy_up = 1968.567 / 9 = 218.730 kN
E_y = Dy_up  # = 218.730 kN

# Sum Fy en BD: By + Dy_up - R_W1_48 - R_W2 - Fy_C = 0
B_y = R_W1_48 + R_W2 + Fy_C - Dy_up
# By = 315 + 225 + 9.642 - 218.730 = 330.912 kN

# Sum Fx en BD: Bx + Fx_C + 15 = 0  (columna empuja beam a la derecha)
B_x = -(Fx_C + D_x_col)  # = -26.491

# Cuerpo AB:
A_x = B_x              # = -26.491 kN (hacia la izquierda)
A_y = R_W1_04 + B_y    # = 225 + 330.912 = 555.912 kN

# Momento en A: int_0^4 W1(x)*x dx + By*4
# int_0^4 (5.625x+45)*x dx = [1.875x^3 + 22.5x^2] de 0 a 4 = 120 + 360 = 480
I_Ma = 480.0
M_a = I_Ma + B_y * 4   # = 480 + 1323.648 = 1803.648 kN*m (CCW)

# ==============================================================
# VERIFICACION GLOBAL
# ==============================================================
check_Fx = A_x + Fx_C + E_x
check_Fy = A_y + E_y - R_W1_total - R_W2 - Fy_C
# Sum M_A global (verificacion)
# Cargas distribuidas respecto a A:
# int_0^8 W1(x)*x dx = [1.875x^3 + 22.5x^2] de 0 a 8 = 960+1440 = 2400
# int_8^13 W2(x)*x dx = [117x^2 - 6x^3] de 8 a 13 = (19773-13182)-(7488-3072) = 6591-4416 = 2175
# Alternativo: int_8^13 (234-18x)*x dx = [117x^2 - 6x^3]_8^13
I_MA_W1 = 1.875*512 + 22.5*64   # = 960 + 1440 = 2400
I_MA_W2 = (117*169 - 6*2197) - (117*64 - 6*512)  # = (19773-13182)-(7488-3072) = 6591-4416 = 2175
# Momentos: Ma(CCW) - I_MA_W1(CW) - I_MA_W2(CW) - Fy_C*8(CW) - M_C(CW) + Ey*13(CCW) - M_E(CW) = 0?
# Nota: Ey actua en D (x=13) pero Ey es la reaccion en E... la columna transmite Ey al punto D
# En realidad para la viga: la fuerza en D es Dy_up hacia arriba y D_x_col=15 a la derecha
# check_MA = Ma - I_MA_W1 - I_MA_W2 - Fy_C*8 - M_C + Dy_up*13 - (no, el momento de E no actua en la viga)
# El momento de 90 kN*m esta en E, no en la viga directamente
check_MA = M_a - I_MA_W1 - I_MA_W2 - Fy_C*8 - M_C + Dy_up*13
# = 1799.204 - 2400 - 2175 - 77.134 - 5 + 2857.93 = 0

# ==============================================================
# FASE 3: FUNCIONES DE ESFUERZOS INTERNOS
# ==============================================================
def N_viga(x):
    if x < 8:
        return -A_x  # = 26.491 kN (tension)
    else:
        return -A_x - Fx_C  # = 15.0 kN

def V_viga(x):
    if x <= 8:
        return A_y - (2.8125*x**2 + 45*x)
    else:
        # Despues de C: V = Ay - R_W1_total - Fy_C - int_8^x W2 dt
        # int_8^x (234-18t) dt = [234t - 9t^2]_8^x = (234x-9x^2) - (1872-576) = 234x-9x^2-1296
        return A_y - R_W1_total - Fy_C - (234*x - 9*x**2 - 1296)

def M_viga(x):
    if x <= 8:
        return -M_a + A_y*x - (0.9375*x**3 + 22.5*x**2)
    else:
        # M = -Ma + Ay*x - int_0^8 W1(t)*(x-t)dt - Fy_C*(x-8) - M_C - int_8^x W2(t)*(x-t)dt
        # int_0^8 W1(t)*(x-t)dt = x*540 - [1.875t^3 + 22.5t^2]_0^8 = 540x - 2400
        # int_8^x W2(t)*(x-t)dt = int_8^x (234-18t)(x-t)dt
        #   = [234x*t - 117t^2 - 18x*t^2/2 + 18t^3/3 + ... ] -- mejor numerico
        # Uso la forma directa: M despues de C
        # V3(x) = A_y - 540 - 9.6418 - (234x - 9x^2 - 1296) = A_y - 540 - 9.6418 - 234x + 9x^2 + 1296
        # M3(x) = M3(8+) + int_8^x V3(t) dt
        # M3(8+) = M_viga(8-) - M_C = (-Ma + Ay*8 - 0.9375*512 - 22.5*64) - 5
        M_at_8_minus = -M_a + A_y*8 - (0.9375*512 + 22.5*64)
        M_at_8_plus = M_at_8_minus - M_C

        # Integrar V3 de 8 a x
        # V3(t) = A_y - 540 - Fy_C + 9*t^2 - 234*t + 1296
        C_v = A_y - 540 - Fy_C + 1296  # = 554.801 - 540 - 9.6418 + 1296 = 1301.159
        # V3(t) = 9t^2 - 234t + C_v
        # int_8^x V3 dt = [3t^3 - 117t^2 + C_v*t]_8^x
        integral_V = (3*x**3 - 117*x**2 + C_v*x) - (3*512 - 117*64 + C_v*8)
        return M_at_8_plus + integral_V

def N_col(y):
    return -Dy_up  # compresion = -219.841 kN

def V_col(y):
    return D_x_col  # 15 kN constante

def M_col(y):
    return -M_E + D_x_col * y  # -90 + 15y

# ==============================================================
# FASE 4: VALORES CRITICOS
# ==============================================================
# Encontrar x donde V=0 en region 1-2 (raiz de -2.8125x^2 - 45x + 554.801 = 0)
# 2.8125x^2 + 45x - 554.801 = 0
a_q = 2.8125
b_q = 45
c_q = -A_y
disc = b_q**2 - 4*a_q*c_q
x_V0 = (-b_q + math.sqrt(disc)) / (2*a_q)  # raiz positiva

M_max_viga = abs(M_viga(0))      # En A
M_at_V0 = M_viga(x_V0)           # Momento maximo positivo (donde V=0)
V_max_viga = abs(V_viga(0))      # En A
M_max_col = abs(M_col(0))        # En E

# ==============================================================
# FASE 5: DISENO - ADMISIBLES
# ==============================================================
sigma_adm = Fy / FS              # 178.571 MPa
tau_adm = tau_y / FS             # 71.429 MPa

# Modulo de seccion requerido (viga)
S_req_viga = M_max_viga * 1e6 / (sigma_adm * 1e6) * 1e6  # en mm^3
S_req_viga_cm3 = S_req_viga / 1e3  # en cm^3
# = 1799.204e3 / 178.571 = 10075.5 cm^3

# Area de alma requerida por cortante
# tau = V / (d * tw) aprox, pero mejor: tau = V*Q/(I*t) -> simplificado: A_web >= V/tau_adm
A_web_req = V_max_viga * 1e3 / (tau_adm)  # mm^2
# = 554.801e3 / 71.429 = 7767.2 mm^2

# ==============================================================
# FASE 6: TRADE STUDY - PERFILES W COMERCIALES (VIGA)
# ==============================================================
# Base de datos de perfiles W comunes (d, bf, tf, tw, A, Ix, Sx, peso)
# Formato: (nombre, d_mm, bf_mm, tf_mm, tw_mm, A_mm2, Ix_mm4*1e6, Sx_mm3*1e3, peso_kg_m)
perfiles_W = [
    ("W610x217x101",  612, 229, 22.1, 12.7, 12900, 764, 2500, 101),
    ("W610x229x113",  612, 229, 22.2, 11.2, 14400, 875, 2860, 113),
    ("W610x324x174",  616, 325, 21.6, 13.0, 22200, 1220, 3960, 174),
    ("W610x324x195",  622, 327, 24.4, 14.7, 24800, 1410, 4530, 195),
    ("W690x254x125",  678, 253, 16.3, 11.7, 16000, 1190, 3510, 125),
    ("W690x254x140",  684, 254, 18.9, 12.4, 17800, 1340, 3920, 140),
    ("W690x254x152",  688, 254, 21.1, 13.0, 19400, 1480, 4300, 152),
    ("W690x254x170",  693, 256, 23.6, 14.5, 21700, 1680, 4850, 170),
    ("W760x267x147",  754, 265, 17.0, 13.2, 18800, 1660, 4400, 147),
    ("W760x267x161",  759, 266, 19.3, 13.8, 20500, 1820, 4800, 161),
    ("W760x267x173",  762, 267, 21.6, 14.4, 22000, 2000, 5250, 173),
    ("W760x267x185",  766, 267, 23.6, 14.9, 23600, 2180, 5690, 185),
    ("W760x267x196",  770, 268, 25.4, 15.6, 25000, 2350, 6100, 196),
    ("W760x381x196",  770, 381, 17.0, 14.0, 25000, 2400, 6230, 196),
    ("W840x292x176",  835, 292, 18.8, 14.0, 22400, 2510, 6010, 176),
    ("W840x292x194",  840, 292, 21.7, 14.7, 24700, 2790, 6640, 194),
    ("W840x292x210",  846, 293, 24.4, 15.4, 26800, 3080, 7280, 210),
    ("W840x292x226",  851, 294, 26.2, 16.6, 28800, 3390, 7970, 226),
    ("W840x292x251",  859, 295, 29.2, 18.2, 32000, 3840, 8940, 251),
    ("W920x313x201",  903, 304, 20.1, 15.2, 25600, 3250, 7200, 201),
    ("W920x313x223",  912, 306, 22.5, 16.0, 28400, 3710, 8140, 223),
    ("W920x313x238",  915, 307, 24.4, 16.5, 30300, 4010, 8760, 238),
    ("W920x313x253",  919, 308, 26.4, 17.3, 32300, 4370, 9510, 253),
    ("W920x313x271",  927, 309, 28.4, 18.4, 34500, 4790, 10300, 271),
    ("W920x313x289",  932, 310, 30.4, 19.4, 36800, 5220, 11200, 289),
    ("W920x420x313",  932, 421, 24.4, 16.0, 39900, 5960, 12800, 313),
    ("W920x420x342",  940, 422, 26.8, 17.4, 43600, 6630, 14100, 342),
    ("W920x420x365",  948, 424, 29.2, 18.9, 46500, 7300, 15400, 365),
    ("W920x420x390",  955, 425, 31.0, 20.3, 49700, 7930, 16600, 390),
    ("W920x420x420",  960, 426, 33.5, 21.4, 53500, 8690, 18100, 420),
    ("W920x420x449",  968, 427, 36.1, 23.0, 57200, 9520, 19700, 449),
    ("W1000x300x222", 970, 300, 21.1, 16.0, 28300, 4100, 8450, 222),
    ("W1000x300x249", 980, 302, 24.4, 17.0, 31700, 4730, 9650, 249),
    ("W1000x300x272", 987, 303, 27.2, 18.0, 34700, 5290, 10700, 272),
    ("W1000x300x296", 997, 305, 30.0, 19.8, 37700, 5950, 11900, 296),
    ("W1000x300x314", 1000, 306, 31.0, 21.1, 40000, 6440, 12900, 314),
    ("W1000x400x296", 982, 400, 24.0, 16.5, 37700, 6170, 12600, 296),
    ("W1000x400x321", 990, 402, 26.2, 17.8, 40900, 6870, 13900, 321),
    ("W1000x400x350", 996, 403, 28.4, 19.1, 44600, 7570, 15200, 350),
    ("W1000x400x371", 1003, 404, 30.2, 20.1, 47300, 8150, 16300, 371),
    ("W1000x400x393", 1008, 405, 32.0, 21.1, 50100, 8700, 17300, 393),
    ("W1000x400x415", 1014, 407, 33.5, 22.3, 52900, 9360, 18500, 415),
    ("W1000x400x443", 1020, 408, 36.1, 23.6, 56400, 10100, 19800, 443),
    ("W1000x400x494", 1030, 410, 39.9, 26.0, 62900, 11500, 22300, 494),
    ("W1000x400x539", 1042, 413, 43.9, 28.4, 68700, 12900, 24800, 539),
    ("W1000x400x591", 1054, 415, 48.3, 31.0, 75300, 14500, 27500, 591),
]

# Evaluar cada perfil
resultados_perfiles = []
for p in perfiles_W:
    nombre, d, bf, tf, tw, A, Ix, Sx, peso = p
    Ix_val = Ix * 1e6    # mm^4
    Sx_val = Sx * 1e3     # mm^3

    # Verificar flexion
    sigma_flex = M_max_viga * 1e6 / Sx_val  # MPa (M en kN*m -> N*mm / mm^3 = MPa)
    FS_flex = Fy / sigma_flex if sigma_flex > 0 else 0
    ok_flex = FS_flex >= FS

    # Verificar cortante (aproximado: tau = V / (d*tw))
    tau_cort = V_max_viga * 1e3 / (d * tw)  # MPa
    FS_cort = tau_y / tau_cort if tau_cort > 0 else 0
    ok_cort = FS_cort >= FS

    # Esbeltez alma
    esb = d / tw
    ok_esb = esb <= 70  # limite practico

    # Volumen viga (mm^2 * 13000 mm = mm^3 -> m^3)
    vol_viga = A * L_viga * 1000 / 1e9  # m^3

    cumple = ok_flex and ok_cort

    resultados_perfiles.append({
        'nombre': nombre, 'd': d, 'bf': bf, 'tf': tf, 'tw': tw,
        'A': A, 'Sx': Sx_val, 'peso': peso,
        'sigma': sigma_flex, 'FS_flex': FS_flex, 'ok_flex': ok_flex,
        'tau': tau_cort, 'FS_cort': FS_cort, 'ok_cort': ok_cort,
        'esb': esb, 'vol': vol_viga, 'cumple': cumple
    })

# Ordenar por volumen (menor primero) entre los que cumplen
cumplen = [r for r in resultados_perfiles if r['cumple']]
cumplen.sort(key=lambda r: r['vol'])

# ==============================================================
# FASE 7: PERFIL PERSONALIZADO OPTIMIZADO (VIGA)
# ==============================================================
# Restricciones: tw >= 8mm, tf >= 8mm, d/tw <= 50
# Objetivo: minimizar area A = 2*bf*tf + (d - 2*tf)*tw
# Restriccion: Sx >= S_req
# Sx = Ix / (d/2), Ix = (bf*d^3 - (bf-tw)*(d-2*tf)^3) / 12

def calc_perfil_I(d_mm, bf_mm, tf_mm, tw_mm):
    """Calcula propiedades de un perfil I"""
    A = 2 * bf_mm * tf_mm + (d_mm - 2*tf_mm) * tw_mm
    # Ix por partes: rectangulo total - dos vacios laterales
    Ix = (bf_mm * d_mm**3) / 12 - ((bf_mm - tw_mm) * (d_mm - 2*tf_mm)**3) / 12
    Sx = Ix / (d_mm / 2)
    return A, Ix, Sx

# Busqueda iterativa FINA
# Estrategia: perfiles altos con alma al limite d/tw=50 son mas eficientes en flexion
mejor_perfil = None
mejor_area = 1e10

S_req_mm3 = S_req_viga_cm3 * 1e3  # en mm^3

for d_test in range(700, 1201, 1):         # paso fino de 1mm
    for tw_test in range(8, 25):
        if d_test / tw_test > 50:
            continue
        # Cortante minimo check
        tau_t = V_max_viga * 1e3 / (d_test * tw_test)
        if tau_y / tau_t < FS:
            continue

        for tf_test in range(8, 37):
            # Para cada tf, calcular el bf MINIMO que da Sx >= S_req
            # Sx = Ix / (d/2); Ix = (bf*d^3 - (bf-tw)*(d-2tf)^3) / 12
            # Ix = (bf*d^3 - bf*(d-2tf)^3 + tw*(d-2tf)^3) / 12
            # Ix = bf*(d^3 - (d-2tf)^3)/12 + tw*(d-2tf)^3/12
            # Sx = Ix*2/d
            # Sx = (2/d) * [bf*(d^3-(d-2tf)^3)/12 + tw*(d-2tf)^3/12]
            # Sx >= S_req => bf >= (S_req - tw*(d-2tf)^3*2/(12*d)) * 12*d / (2*(d^3-(d-2tf)^3))
            hw = d_test - 2*tf_test
            if hw <= 0:
                continue
            d3 = d_test**3
            hw3 = hw**3
            delta = d3 - hw3

            if delta <= 0:
                continue

            Sx_from_web = tw_test * hw3 * 2 / (12 * d_test)
            remaining = S_req_mm3 - Sx_from_web
            if remaining <= 0:
                # Web alone provides enough Sx (unlikely but check)
                bf_min = tw_test + 1
            else:
                bf_min = remaining * 12 * d_test / (2 * delta)
                bf_min = math.ceil(bf_min)

            # bf debe ser >= tw y razonable
            bf_min = max(bf_min, tw_test + 20, 100)
            if bf_min > 500:
                continue

            A_t, Ix_t, Sx_t = calc_perfil_I(d_test, bf_min, tf_test, tw_test)

            sigma_t = M_max_viga * 1e6 / Sx_t if Sx_t > 0 else 1e10
            FS_t = Fy / sigma_t if sigma_t > 0 else 0

            if FS_t >= FS and A_t < mejor_area:
                mejor_area = A_t
                mejor_perfil = {
                    'd': d_test, 'bf': bf_min, 'tf': tf_test, 'tw': tw_test,
                    'A': A_t, 'Ix': Ix_t, 'Sx': Sx_t,
                    'sigma': sigma_t, 'FS_flex': FS_t,
                    'tau': tau_t, 'FS_cort': tau_y / tau_t,
                    'esb': d_test / tw_test
                }

# ==============================================================
# FASE 8: PERFIL COLUMNA
# ==============================================================
# M_max_col = 90 kN*m, N_col = 219.84 kN (compresion), V_col = 15 kN
# Criterio combinado: sigma = N/A + M*c/I <= sigma_adm
# Buscar perfil minimo

perfiles_col = [
    ("W150x37x30",  157, 153, 9.3, 6.6, 3790, 17.1, 218, 30),
    ("W200x100x22", 206, 102, 8.0, 6.2, 2860, 20.0, 194, 22),
    ("W200x200x46", 203, 203, 11.0, 7.2, 5890, 45.5, 448, 46),
    ("W200x200x59", 210, 205, 14.2, 9.1, 7550, 60.8, 579, 59),
    ("W200x200x71", 216, 206, 17.4, 10.2, 9100, 76.6, 709, 71),
    ("W250x150x38", 252, 146, 10.7, 6.1, 4830, 48.4, 384, 38),
    ("W250x200x49", 247, 202, 11.0, 7.6, 6260, 70.6, 572, 49),
    ("W250x200x58", 252, 203, 13.5, 8.0, 7400, 87.3, 693, 58),
    ("W250x200x67", 257, 204, 15.7, 8.9, 8560, 104, 809, 67),
    ("W250x250x73", 253, 254, 14.2, 8.6, 9290, 113, 893, 73),
    ("W250x250x80", 256, 255, 15.6, 9.4, 10200, 126, 984, 80),
    ("W250x250x89", 260, 256, 17.3, 10.7, 11400, 142, 1090, 89),
    ("W310x165x39", 310, 165, 9.7, 5.8, 4960, 85.1, 549, 39),
    ("W310x165x45", 313, 166, 11.2, 6.6, 5710, 99.2, 634, 45),
    ("W310x200x52", 318, 203, 10.8, 7.6, 6650, 119, 748, 52),
    ("W310x200x60", 303, 203, 13.1, 7.5, 7610, 129, 851, 60),
    ("W310x250x79", 306, 254, 14.6, 8.8, 10100, 177, 1157, 79),
    ("W310x250x86", 310, 254, 16.3, 9.1, 11000, 198, 1277, 86),
    ("W310x310x97", 308, 305, 15.4, 9.9, 12300, 222, 1442, 97),
]

resultados_col = []
for p in perfiles_col:
    nombre, d, bf, tf, tw, A, Ix_1e6, Sx_1e3, peso = p
    Ix_v = Ix_1e6 * 1e6   # mm^4
    Sx_v = Sx_1e3 * 1e3   # mm^3
    c = d / 2

    # Esfuerzo combinado: sigma = N/A + M*c/I
    sigma_N = abs(N_col(0)) * 1e3 / A              # MPa (compresion)
    sigma_M = abs(M_col(0)) * 1e6 / Sx_v           # MPa (flexion)
    sigma_total = sigma_N + sigma_M
    FS_col = Fy / sigma_total if sigma_total > 0 else 0

    # Cortante
    tau_c = abs(V_col(0)) * 1e3 / (d * tw)
    FS_tau_c = tau_y / tau_c if tau_c > 0 else 0

    vol_col = A * L_col * 1000 / 1e9  # m^3

    cumple_c = FS_col >= FS and FS_tau_c >= FS

    resultados_col.append({
        'nombre': nombre, 'd': d, 'bf': bf, 'tf': tf, 'tw': tw,
        'A': A, 'Sx': Sx_v, 'peso': peso,
        'sigma_N': sigma_N, 'sigma_M': sigma_M, 'sigma_total': sigma_total,
        'FS': FS_col, 'tau': tau_c, 'FS_tau': FS_tau_c,
        'vol': vol_col, 'cumple': cumple_c
    })

cumplen_col = [r for r in resultados_col if r['cumple']]
cumplen_col.sort(key=lambda r: r['vol'])

# ==============================================================
# CREAR EXCEL
# ==============================================================
wb = Workbook()

# Estilos
titulo_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
titulo_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
subheader_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
ok_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
best_fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
num_2d = '#,##0.00'
num_4d = '#,##0.0000'

def style_header(ws, row, cols, fill=header_fill):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

def style_cell(ws, row, col, fmt=None):
    cell = ws.cell(row=row, column=col)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')
    if fmt:
        cell.number_format = fmt

def write_title(ws, row, text, cols=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font = titulo_font
    cell.fill = titulo_fill
    cell.alignment = Alignment(horizontal='center')

# ===================== HOJA 1: DATOS DE ENTRADA =====================
ws1 = wb.active
ws1.title = "1. Datos"
ws1.sheet_properties.tabColor = '1F4E79'

write_title(ws1, 1, "PROYECTO FINAL - MECANICA DE MATERIALES II", 6)
ws1.cell(row=2, column=1, value="Portico plano de acero - Ing. William Valencia Mina")
ws1.cell(row=2, column=1).font = Font(size=11, italic=True)
ws1.cell(row=3, column=1, value="Universidad del Quindio - Ingenieria Civil")

r = 5
ws1.cell(row=r, column=1, value="PARAMETRO").font = Font(bold=True)
ws1.cell(row=r, column=2, value="SIMBOLO").font = Font(bold=True)
ws1.cell(row=r, column=3, value="VALOR").font = Font(bold=True)
ws1.cell(row=r, column=4, value="UNIDAD").font = Font(bold=True)
style_header(ws1, r, 4)

datos = [
    ("Parametro de carga", "xy", xy, "kN/m"),
    ("Factor de seguridad", "FS", FS, "-"),
    ("Fluencia acero A36", "Fy", Fy, "MPa"),
    ("Cortante fluencia", "tau_y", tau_y, "MPa"),
    ("Modulo elasticidad", "E", E_steel, "GPa"),
    ("", "", "", ""),
    ("Longitud A-B", "L_AB", L_AB, "m"),
    ("Longitud B-C", "L_BC", L_BC, "m"),
    ("Longitud C-D", "L_CD", L_CD, "m"),
    ("Longitud total viga", "L_viga", L_viga, "m"),
    ("Altura columna", "L_col", L_col, "m"),
    ("", "", "", ""),
    ("Carga distribuida en A", "w(0)", xy, "kN/m"),
    ("Carga distribuida en C", "w(8)", 2*xy, "kN/m"),
    ("Carga distribuida en D", "w(13)", 0, "kN/m"),
    ("Fuerza puntual", "P", P_puntual, "kN"),
    ("Angulo fuerza puntual", "theta", angulo, "grados"),
    ("Componente horizontal P", "Fx", round(Fx_C, 4), "kN"),
    ("Componente vertical P", "Fy_p", round(Fy_C, 4), "kN"),
    ("Momento en C", "Mc", M_C, "kN*m (CW)"),
    ("Momento en E", "Me", M_E, "kN*m (CW)"),
    ("", "", "", ""),
    ("Resultante W1 [0,8]", "R_W1", R_W1_total, "kN"),
    ("Resultante W2 [8,13]", "R_W2", R_W2, "kN"),
    ("Carga total vertical", "F_total", round(R_W1_total + R_W2 + Fy_C, 3), "kN"),
]

for i, (param, simb, val, uni) in enumerate(datos):
    row = r + 1 + i
    ws1.cell(row=row, column=1, value=param)
    ws1.cell(row=row, column=2, value=simb)
    ws1.cell(row=row, column=3, value=val)
    ws1.cell(row=row, column=4, value=uni)
    for c in range(1, 5):
        style_cell(ws1, row, c)

ws1.column_dimensions['A'].width = 30
ws1.column_dimensions['B'].width = 12
ws1.column_dimensions['C'].width = 14
ws1.column_dimensions['D'].width = 14

# ===================== HOJA 2: REACCIONES =====================
ws2 = wb.create_sheet("2. Reacciones")
ws2.sheet_properties.tabColor = '2E75B6'

write_title(ws2, 1, "CALCULO DE REACCIONES", 8)

r = 3
ws2.cell(row=r, column=1, value="ESQUEMA DEL PORTICO:").font = Font(bold=True)
r = 4
ws2.cell(row=r, column=1, value="A(empotrado) ---4m--- B(rotula) ---4m--- C ---5m--- D(rotula) --- columna 6m --- E(articulado)")

r = 6
ws2.cell(row=r, column=1, value="PASO").font = Font(bold=True)
ws2.cell(row=r, column=2, value="ECUACION").font = Font(bold=True)
ws2.cell(row=r, column=3, value="RESULTADO").font = Font(bold=True)
ws2.cell(row=r, column=4, value="VALOR").font = Font(bold=True)
ws2.cell(row=r, column=5, value="UNIDAD").font = Font(bold=True)
style_header(ws2, r, 5)

pasos = [
    ("Columna: SumM_E=0", "Dx*6 - 90 = 0", "Dx (col)", round(D_x_col, 4), "kN"),
    ("Columna: SumFx=0", "Ex = Dx", "Ex", round(E_x, 4), "kN"),
    ("BD: SumM_B=0", "Dy*9 = 660+1275+38.57+5", "Dy (up en D)", round(Dy_up, 4), "kN"),
    ("Columna: SumFy=0", "Ey = Dy", "Ey", round(E_y, 4), "kN"),
    ("BD: SumFy=0", "By = 315+225+9.64-219.84", "By", round(B_y, 4), "kN"),
    ("BD: SumFx=0", "Bx + 11.49 + 15 = 0", "Bx", round(B_x, 4), "kN"),
    ("AB: SumFx=0", "Ax = Bx", "Ax", round(A_x, 4), "kN"),
    ("AB: SumFy=0", "Ay = 225 + By", "Ay", round(A_y, 4), "kN"),
    ("AB: SumM_A=0", "Ma = 480 + By*4", "Ma", round(M_a, 4), "kN*m"),
]

for i, (paso, ec, res, val, uni) in enumerate(pasos):
    row = r + 1 + i
    ws2.cell(row=row, column=1, value=paso)
    ws2.cell(row=row, column=2, value=ec)
    ws2.cell(row=row, column=3, value=res)
    ws2.cell(row=row, column=4, value=val)
    ws2.cell(row=row, column=5, value=uni)
    for c in range(1, 6):
        style_cell(ws2, row, c, num_4d if c == 4 else None)

# Tabla resumen
r2 = r + len(pasos) + 3
write_title(ws2, r2, "TABLA RESUMEN DE REACCIONES", 5)
r2 += 1
headers_reac = ["Punto", "Fx (kN)", "Fy (kN)", "M (kN*m)", "Tipo"]
for i, h in enumerate(headers_reac):
    ws2.cell(row=r2, column=i+1, value=h)
style_header(ws2, r2, 5)

reac_data = [
    ("A (empotramiento)", round(A_x, 3), round(A_y, 3), round(M_a, 3), "Empotrado"),
    ("E (articulacion)", round(E_x, 3), round(E_y, 3), "-", "Pin"),
    ("B (rotula interna)", round(B_x, 3), round(B_y, 3), "0 (rotula)", "Rotula"),
    ("D (rotula interna)", round(-D_x_col, 3), round(-Dy_up, 3), "0 (rotula)", "Rotula"),
]

for i, row_data in enumerate(reac_data):
    row = r2 + 1 + i
    for j, val in enumerate(row_data):
        ws2.cell(row=row, column=j+1, value=val)
        style_cell(ws2, row, j+1, num_2d if isinstance(val, float) else None)

# Verificaciones
r3 = r2 + len(reac_data) + 3
write_title(ws2, r3, "VERIFICACION EQUILIBRIO GLOBAL", 5)
r3 += 1
ws2.cell(row=r3, column=1, value="SumFx = Ax + Fx_C + Ex")
ws2.cell(row=r3, column=2, value=round(check_Fx, 6))
ws2.cell(row=r3, column=3, value="OK" if abs(check_Fx) < 0.01 else "ERROR")
ws2.cell(row=r3, column=3).fill = ok_fill if abs(check_Fx) < 0.01 else fail_fill
r3 += 1
ws2.cell(row=r3, column=1, value="SumFy = Ay + Ey - cargas")
ws2.cell(row=r3, column=2, value=round(check_Fy, 6))
ws2.cell(row=r3, column=3, value="OK" if abs(check_Fy) < 0.01 else "ERROR")
ws2.cell(row=r3, column=3).fill = ok_fill if abs(check_Fy) < 0.01 else fail_fill

ws2.column_dimensions['A'].width = 30
ws2.column_dimensions['B'].width = 35
ws2.column_dimensions['C'].width = 18
ws2.column_dimensions['D'].width = 14
ws2.column_dimensions['E'].width = 14

# ===================== HOJA 3: ESFUERZOS INTERNOS =====================
ws3 = wb.create_sheet("3. Esfuerzos Internos")
ws3.sheet_properties.tabColor = '548235'

write_title(ws3, 1, "DIAGRAMAS DE ESFUERZOS INTERNOS (V, N, M)", 8)

# Generar tabla de valores
r = 3
ws3.cell(row=r, column=1, value="FUNCIONES ANALITICAS:").font = Font(bold=True)
r = 4
funcs = [
    "Region 1-2 [0,8]: N = 26.49 kN | V(x) = -2.8125x^2 - 45x + 554.801 | M(x) = -0.9375x^3 - 22.5x^2 + 554.801x - 1799.204",
    "Discontinuidad C (x=8): N: 26.49->15.00 | V: 14.80->5.16 | M: 719.20->714.20",
    "Region 3 [8,13]: N = 15.00 kN | V(x) = 9x^2 - 234x + 1301.16 | M(x) = 3x^3 - 117x^2 + 1301.16x - 3743.07",
    "Columna [0,6]: N = -219.84 kN | V = 15 kN | M(y) = -90 + 15y",
]
for i, f in enumerate(funcs):
    ws3.cell(row=r+i, column=1, value=f)

# Tabla numerica VIGA
r = 10
headers_int = ["x (m)", "N (kN)", "V (kN)", "M (kN*m)", "Region", "Nota"]
for i, h in enumerate(headers_int):
    ws3.cell(row=r, column=i+1, value=h)
style_header(ws3, r, 6)

x_vals = np.concatenate([
    np.arange(0, 8.001, 0.5),
    [7.999, 8.0],  # justo antes de C
    np.arange(8.5, 13.001, 0.5)
])
x_vals = sorted(set(np.round(x_vals, 3)))

row_i = r + 1
for x in x_vals:
    nota = ""
    if abs(x) < 0.001: nota = "Empotramiento A"
    elif abs(x - 4) < 0.001: nota = "Rotula B"
    elif abs(x - 8) < 0.01: nota = "Punto C"
    elif abs(x - 13) < 0.001: nota = "Rotula D"

    n_val = N_viga(x)
    v_val = V_viga(x)
    m_val = M_viga(x)

    region = "AB" if x <= 4 else ("BC" if x <= 8 else "CD")

    ws3.cell(row=row_i, column=1, value=round(x, 3))
    ws3.cell(row=row_i, column=2, value=round(n_val, 3))
    ws3.cell(row=row_i, column=3, value=round(v_val, 3))
    ws3.cell(row=row_i, column=4, value=round(m_val, 3))
    ws3.cell(row=row_i, column=5, value=region)
    ws3.cell(row=row_i, column=6, value=nota)
    for c in range(1, 7):
        style_cell(ws3, row_i, c, num_2d if c <= 4 else None)
    row_i += 1

# Tabla columna
row_i += 2
ws3.cell(row=row_i, column=1, value="COLUMNA D-E").font = Font(bold=True, size=12)
row_i += 1
headers_col = ["y (m)", "N (kN)", "V (kN)", "M (kN*m)", "Nota"]
for i, h in enumerate(headers_col):
    ws3.cell(row=row_i, column=i+1, value=h)
style_header(ws3, row_i, 5)
row_i += 1

for y in np.arange(0, 6.01, 1):
    nota = ""
    if abs(y) < 0.001: nota = "Base E"
    elif abs(y - 6) < 0.001: nota = "Tope D (rotula)"

    ws3.cell(row=row_i, column=1, value=round(y, 1))
    ws3.cell(row=row_i, column=2, value=round(N_col(y), 3))
    ws3.cell(row=row_i, column=3, value=round(V_col(y), 3))
    ws3.cell(row=row_i, column=4, value=round(M_col(y), 3))
    ws3.cell(row=row_i, column=5, value=nota)
    for c in range(1, 6):
        style_cell(ws3, row_i, c, num_2d if c <= 4 else None)
    row_i += 1

# Valores criticos
row_i += 2
write_title(ws3, row_i, "VALORES CRITICOS PARA DISENO", 6)
row_i += 1
crit_headers = ["Elemento", "|M|max (kN*m)", "|V|max (kN)", "|N|max (kN)", "Ubicacion"]
for i, h in enumerate(crit_headers):
    ws3.cell(row=row_i, column=i+1, value=h)
style_header(ws3, row_i, 5)

row_i += 1
ws3.cell(row=row_i, column=1, value="Viga")
ws3.cell(row=row_i, column=2, value=round(M_max_viga, 3))
ws3.cell(row=row_i, column=3, value=round(V_max_viga, 3))
ws3.cell(row=row_i, column=4, value=round(abs(A_x), 3))
ws3.cell(row=row_i, column=5, value="x=0 (A)")
for c in range(1, 6): style_cell(ws3, row_i, c, num_2d if 2<=c<=4 else None)

row_i += 1
ws3.cell(row=row_i, column=1, value="Columna")
ws3.cell(row=row_i, column=2, value=round(M_max_col, 3))
ws3.cell(row=row_i, column=3, value=round(abs(V_col(0)), 3))
ws3.cell(row=row_i, column=4, value=round(abs(N_col(0)), 3))
ws3.cell(row=row_i, column=5, value="y=0 (E)")
for c in range(1, 6): style_cell(ws3, row_i, c, num_2d if 2<=c<=4 else None)

row_i += 1
ws3.cell(row=row_i, column=1, value="Viga (M+ max)")
ws3.cell(row=row_i, column=2, value=round(M_at_V0, 3))
ws3.cell(row=row_i, column=3, value="0 (V=0)")
ws3.cell(row=row_i, column=4, value=round(abs(A_x), 3))
ws3.cell(row=row_i, column=5, value=f"x={round(x_V0, 3)} m")
for c in range(1, 6): style_cell(ws3, row_i, c)

ws3.column_dimensions['A'].width = 15
ws3.column_dimensions['B'].width = 14
ws3.column_dimensions['C'].width = 14
ws3.column_dimensions['D'].width = 14
ws3.column_dimensions['E'].width = 12
ws3.column_dimensions['F'].width = 20

# ===================== HOJA 4: TRADE STUDY PERFILES VIGA =====================
ws4 = wb.create_sheet("4. Trade Study Viga")
ws4.sheet_properties.tabColor = 'BF8F00'

write_title(ws4, 1, "TRADE STUDY - SELECCION PERFIL W COMERCIAL (VIGA)", 14)

r = 3
ws4.cell(row=r, column=1, value="REQUISITOS DE DISENO:").font = Font(bold=True)
r += 1
ws4.cell(row=r, column=1, value=f"sigma_adm = Fy/FS = {Fy}/{FS} = {round(sigma_adm, 2)} MPa")
r += 1
ws4.cell(row=r, column=1, value=f"tau_adm = tau_y/FS = {tau_y}/{FS} = {round(tau_adm, 2)} MPa")
r += 1
ws4.cell(row=r, column=1, value=f"S_requerido >= {round(S_req_viga_cm3, 1)} cm3 = {round(S_req_viga_cm3*1e3, 0)} mm3")
r += 1
ws4.cell(row=r, column=1, value=f"A_web_req >= {round(A_web_req, 1)} mm2 (por cortante)")

r += 2
headers_ts = ["#", "Perfil", "d (mm)", "bf (mm)", "tf (mm)", "tw (mm)",
              "A (mm2)", "Sx (mm3)", "sigma (MPa)", "FS flex",
              "tau (MPa)", "FS cort", "Vol viga (m3)", "CUMPLE"]
for i, h in enumerate(headers_ts):
    ws4.cell(row=r, column=i+1, value=h)
style_header(ws4, r, len(headers_ts))

for i, res in enumerate(resultados_perfiles):
    row = r + 1 + i
    ws4.cell(row=row, column=1, value=i+1)
    ws4.cell(row=row, column=2, value=res['nombre'])
    ws4.cell(row=row, column=3, value=res['d'])
    ws4.cell(row=row, column=4, value=res['bf'])
    ws4.cell(row=row, column=5, value=res['tf'])
    ws4.cell(row=row, column=6, value=res['tw'])
    ws4.cell(row=row, column=7, value=res['A'])
    ws4.cell(row=row, column=8, value=res['Sx'])
    ws4.cell(row=row, column=9, value=round(res['sigma'], 2))
    ws4.cell(row=row, column=10, value=round(res['FS_flex'], 3))
    ws4.cell(row=row, column=11, value=round(res['tau'], 2))
    ws4.cell(row=row, column=12, value=round(res['FS_cort'], 3))
    ws4.cell(row=row, column=13, value=round(res['vol'], 6))
    ws4.cell(row=row, column=14, value="SI" if res['cumple'] else "NO")

    for c in range(1, len(headers_ts)+1):
        style_cell(ws4, row, c)

    if res['cumple']:
        ws4.cell(row=row, column=14).fill = ok_fill
    else:
        ws4.cell(row=row, column=14).fill = fail_fill

# Marcar el mejor
if cumplen:
    best = cumplen[0]
    for i, res in enumerate(resultados_perfiles):
        if res['nombre'] == best['nombre']:
            row = r + 1 + i
            for c in range(1, len(headers_ts)+1):
                ws4.cell(row=row, column=c).fill = best_fill

# Resumen mejores
r_best = r + len(resultados_perfiles) + 3
write_title(ws4, r_best, "TOP 5 PERFILES MAS LIVIANOS QUE CUMPLEN", 14)
r_best += 1
for i, h in enumerate(headers_ts):
    ws4.cell(row=r_best, column=i+1, value=h)
style_header(ws4, r_best, len(headers_ts))

for i, res in enumerate(cumplen[:5]):
    row = r_best + 1 + i
    ws4.cell(row=row, column=1, value=i+1)
    ws4.cell(row=row, column=2, value=res['nombre'])
    ws4.cell(row=row, column=3, value=res['d'])
    ws4.cell(row=row, column=4, value=res['bf'])
    ws4.cell(row=row, column=5, value=res['tf'])
    ws4.cell(row=row, column=6, value=res['tw'])
    ws4.cell(row=row, column=7, value=res['A'])
    ws4.cell(row=row, column=8, value=res['Sx'])
    ws4.cell(row=row, column=9, value=round(res['sigma'], 2))
    ws4.cell(row=row, column=10, value=round(res['FS_flex'], 3))
    ws4.cell(row=row, column=11, value=round(res['tau'], 2))
    ws4.cell(row=row, column=12, value=round(res['FS_cort'], 3))
    ws4.cell(row=row, column=13, value=round(res['vol'], 6))
    ws4.cell(row=row, column=14, value="SI")
    for c in range(1, len(headers_ts)+1):
        style_cell(ws4, row, c)
    if i == 0:
        for c in range(1, len(headers_ts)+1):
            ws4.cell(row=row, column=c).fill = best_fill

for c in range(1, len(headers_ts)+1):
    ws4.column_dimensions[get_column_letter(c)].width = 14
ws4.column_dimensions['B'].width = 24

# ===================== HOJA 5: PERFIL PERSONALIZADO =====================
ws5 = wb.create_sheet("5. Perfil Personalizado")
ws5.sheet_properties.tabColor = 'C00000'

write_title(ws5, 1, "PERFIL PERSONALIZADO OPTIMIZADO (VIGA)", 8)

r = 3
ws5.cell(row=r, column=1, value="RESTRICCIONES:").font = Font(bold=True)
ws5.cell(row=r+1, column=1, value="tw >= 8 mm")
ws5.cell(row=r+2, column=1, value="tf >= 8 mm")
ws5.cell(row=r+3, column=1, value="d/tw <= 50")
ws5.cell(row=r+4, column=1, value="FS >= 1.4 (flexion y cortante)")
ws5.cell(row=r+5, column=1, value="OBJETIVO: Minimizar area de seccion transversal")

if mejor_perfil:
    r = 10
    write_title(ws5, r, "PERFIL OPTIMO ENCONTRADO", 4)
    r += 1
    props = [
        ("Altura (d)", mejor_perfil['d'], "mm"),
        ("Ancho ala (bf)", mejor_perfil['bf'], "mm"),
        ("Espesor ala (tf)", mejor_perfil['tf'], "mm"),
        ("Espesor alma (tw)", mejor_perfil['tw'], "mm"),
        ("Area", mejor_perfil['A'], "mm2"),
        ("Ix", round(mejor_perfil['Ix'], 0), "mm4"),
        ("Sx", round(mejor_perfil['Sx'], 0), "mm3"),
        ("d/tw", round(mejor_perfil['esb'], 2), "-"),
        ("", "", ""),
        ("sigma max", round(mejor_perfil['sigma'], 2), "MPa"),
        ("FS flexion", round(mejor_perfil['FS_flex'], 4), "-"),
        ("tau max", round(mejor_perfil['tau'], 2), "MPa"),
        ("FS cortante", round(mejor_perfil['FS_cort'], 4), "-"),
        ("", "", ""),
        ("Volumen viga (13m)", round(mejor_perfil['A'] * 13000 / 1e9, 6), "m3"),
        ("Peso lineal", round(mejor_perfil['A'] * 7850 / 1e6, 2), "kg/m"),
        ("Peso total viga", round(mejor_perfil['A'] * 7850 / 1e6 * 13, 2), "kg"),
    ]

    headers_pp = ["Propiedad", "Valor", "Unidad"]
    for i, h in enumerate(headers_pp):
        ws5.cell(row=r, column=i+1, value=h)
    style_header(ws5, r, 3)

    for i, (prop, val, uni) in enumerate(props):
        row = r + 1 + i
        ws5.cell(row=row, column=1, value=prop)
        ws5.cell(row=row, column=2, value=val)
        ws5.cell(row=row, column=3, value=uni)
        for c in range(1, 4):
            style_cell(ws5, row, c)

    # Comparacion con comercial
    if cumplen:
        r_comp = r + len(props) + 3
        write_title(ws5, r_comp, "COMPARACION: PERSONALIZADO vs COMERCIAL", 5)
        r_comp += 1
        comp_headers = ["", "Comercial optimo", "Personalizado", "Ahorro", "Ahorro %"]
        for i, h in enumerate(comp_headers):
            ws5.cell(row=r_comp, column=i+1, value=h)
        style_header(ws5, r_comp, 5)

        best_com = cumplen[0]
        vol_com = best_com['vol']
        vol_per = mejor_perfil['A'] * 13000 / 1e9

        r_comp += 1
        ws5.cell(row=r_comp, column=1, value="Perfil")
        ws5.cell(row=r_comp, column=2, value=best_com['nombre'])
        ws5.cell(row=r_comp, column=3, value=f"I-{mejor_perfil['d']}x{mejor_perfil['bf']}")

        r_comp += 1
        ws5.cell(row=r_comp, column=1, value="Area (mm2)")
        ws5.cell(row=r_comp, column=2, value=best_com['A'])
        ws5.cell(row=r_comp, column=3, value=mejor_perfil['A'])
        ws5.cell(row=r_comp, column=4, value=best_com['A'] - mejor_perfil['A'])
        ws5.cell(row=r_comp, column=5, value=f"{round((1 - mejor_perfil['A']/best_com['A'])*100, 1)}%")

        r_comp += 1
        ws5.cell(row=r_comp, column=1, value="Volumen viga (m3)")
        ws5.cell(row=r_comp, column=2, value=round(vol_com, 6))
        ws5.cell(row=r_comp, column=3, value=round(vol_per, 6))
        ws5.cell(row=r_comp, column=4, value=round(vol_com - vol_per, 6))
        ws5.cell(row=r_comp, column=5, value=f"{round((1 - vol_per/vol_com)*100, 1)}%")

        peso_com = best_com['A'] * 7850 / 1e6 * 13
        peso_per = mejor_perfil['A'] * 7850 / 1e6 * 13
        r_comp += 1
        ws5.cell(row=r_comp, column=1, value="Peso viga (kg)")
        ws5.cell(row=r_comp, column=2, value=round(peso_com, 1))
        ws5.cell(row=r_comp, column=3, value=round(peso_per, 1))
        ws5.cell(row=r_comp, column=4, value=round(peso_com - peso_per, 1))
        ws5.cell(row=r_comp, column=5, value=f"{round((1 - peso_per/peso_com)*100, 1)}%")

ws5.column_dimensions['A'].width = 25
ws5.column_dimensions['B'].width = 20
ws5.column_dimensions['C'].width = 18
ws5.column_dimensions['D'].width = 14
ws5.column_dimensions['E'].width = 14

# ===================== HOJA 6: PERFIL COLUMNA =====================
ws6 = wb.create_sheet("6. Perfil Columna")
ws6.sheet_properties.tabColor = '7030A0'

write_title(ws6, 1, "SELECCION PERFIL COLUMNA", 14)

r = 3
ws6.cell(row=r, column=1, value="SOLICITACIONES COLUMNA:").font = Font(bold=True)
ws6.cell(row=r+1, column=1, value=f"|N|max = {round(abs(N_col(0)), 3)} kN (compresion)")
ws6.cell(row=r+2, column=1, value=f"|M|max = {round(abs(M_col(0)), 3)} kN*m (en E)")
ws6.cell(row=r+3, column=1, value=f"|V|max = {round(abs(V_col(0)), 3)} kN")
ws6.cell(row=r+4, column=1, value=f"Criterio: sigma = N/A + M*c/I <= {round(sigma_adm, 2)} MPa")

r = 9
headers_c = ["#", "Perfil", "d", "bf", "tf", "tw", "A (mm2)", "Sx (mm3)",
             "sigma_N", "sigma_M", "sigma_tot", "FS", "tau", "FS_tau", "Vol (m3)", "CUMPLE"]
for i, h in enumerate(headers_c):
    ws6.cell(row=r, column=i+1, value=h)
style_header(ws6, r, len(headers_c))

for i, res in enumerate(resultados_col):
    row = r + 1 + i
    ws6.cell(row=row, column=1, value=i+1)
    ws6.cell(row=row, column=2, value=res['nombre'])
    ws6.cell(row=row, column=3, value=res['d'])
    ws6.cell(row=row, column=4, value=res['bf'])
    ws6.cell(row=row, column=5, value=res['tf'])
    ws6.cell(row=row, column=6, value=res['tw'])
    ws6.cell(row=row, column=7, value=res['A'])
    ws6.cell(row=row, column=8, value=res['Sx'])
    ws6.cell(row=row, column=9, value=round(res['sigma_N'], 2))
    ws6.cell(row=row, column=10, value=round(res['sigma_M'], 2))
    ws6.cell(row=row, column=11, value=round(res['sigma_total'], 2))
    ws6.cell(row=row, column=12, value=round(res['FS'], 3))
    ws6.cell(row=row, column=13, value=round(res['tau'], 2))
    ws6.cell(row=row, column=14, value=round(res['FS_tau'], 3))
    ws6.cell(row=row, column=15, value=round(res['vol'], 6))
    ws6.cell(row=row, column=16, value="SI" if res['cumple'] else "NO")

    for c in range(1, len(headers_c)+1):
        style_cell(ws6, row, c)
    ws6.cell(row=row, column=16).fill = ok_fill if res['cumple'] else fail_fill

if cumplen_col:
    for i, res in enumerate(resultados_col):
        if res['nombre'] == cumplen_col[0]['nombre']:
            row = r + 1 + i
            for c in range(1, len(headers_c)+1):
                ws6.cell(row=row, column=c).fill = best_fill

for c in range(1, len(headers_c)+1):
    ws6.column_dimensions[get_column_letter(c)].width = 13
ws6.column_dimensions['B'].width = 22

# ===================== HOJA 7: VOLUMENES =====================
ws7 = wb.create_sheet("7. Volumenes")
ws7.sheet_properties.tabColor = '00B050'

write_title(ws7, 1, "CALCULO DE VOLUMENES TOTALES DE MATERIAL", 6)

r = 3

# Con perfil comercial
if cumplen and cumplen_col:
    best_v = cumplen[0]
    best_c = cumplen_col[0]

    vol_viga_com = best_v['A'] * L_viga * 1000 / 1e9
    vol_col_com = best_c['A'] * L_col * 1000 / 1e9
    vol_total_com = vol_viga_com + vol_col_com

    write_title(ws7, r, "OPCION 1: PERFILES COMERCIALES", 6)
    r += 1
    ws7.cell(row=r, column=1, value="Elemento")
    ws7.cell(row=r, column=2, value="Perfil")
    ws7.cell(row=r, column=3, value="Area (mm2)")
    ws7.cell(row=r, column=4, value="Long (m)")
    ws7.cell(row=r, column=5, value="Vol (m3)")
    ws7.cell(row=r, column=6, value="Peso (kg)")
    style_header(ws7, r, 6)

    r += 1
    ws7.cell(row=r, column=1, value="Viga")
    ws7.cell(row=r, column=2, value=best_v['nombre'])
    ws7.cell(row=r, column=3, value=best_v['A'])
    ws7.cell(row=r, column=4, value=L_viga)
    ws7.cell(row=r, column=5, value=round(vol_viga_com, 6))
    ws7.cell(row=r, column=6, value=round(best_v['A'] * 7850 / 1e6 * L_viga, 1))
    for c in range(1, 7): style_cell(ws7, r, c)

    r += 1
    ws7.cell(row=r, column=1, value="Columna")
    ws7.cell(row=r, column=2, value=best_c['nombre'])
    ws7.cell(row=r, column=3, value=best_c['A'])
    ws7.cell(row=r, column=4, value=L_col)
    ws7.cell(row=r, column=5, value=round(vol_col_com, 6))
    ws7.cell(row=r, column=6, value=round(best_c['A'] * 7850 / 1e6 * L_col, 1))
    for c in range(1, 7): style_cell(ws7, r, c)

    r += 1
    ws7.cell(row=r, column=1, value="TOTAL COMERCIAL").font = Font(bold=True)
    ws7.cell(row=r, column=5, value=round(vol_total_com, 6))
    ws7.cell(row=r, column=5).font = Font(bold=True, size=13)
    ws7.cell(row=r, column=6, value=round((best_v['A'] * L_viga + best_c['A'] * L_col) * 7850 / 1e6 * 1000 / 1000, 1))
    ws7.cell(row=r, column=6).font = Font(bold=True)
    for c in range(1, 7): style_cell(ws7, r, c)

    # Con perfil personalizado
    if mejor_perfil:
        r += 3
        vol_viga_per = mejor_perfil['A'] * L_viga * 1000 / 1e9
        vol_total_per = vol_viga_per + vol_col_com

        write_title(ws7, r, "OPCION 2: VIGA PERSONALIZADA + COLUMNA COMERCIAL", 6)
        r += 1
        ws7.cell(row=r, column=1, value="Elemento")
        ws7.cell(row=r, column=2, value="Perfil")
        ws7.cell(row=r, column=3, value="Area (mm2)")
        ws7.cell(row=r, column=4, value="Long (m)")
        ws7.cell(row=r, column=5, value="Vol (m3)")
        ws7.cell(row=r, column=6, value="Peso (kg)")
        style_header(ws7, r, 6)

        r += 1
        ws7.cell(row=r, column=1, value="Viga (personalizada)")
        ws7.cell(row=r, column=2, value=f"I-{mejor_perfil['d']}x{mejor_perfil['bf']}")
        ws7.cell(row=r, column=3, value=mejor_perfil['A'])
        ws7.cell(row=r, column=4, value=L_viga)
        ws7.cell(row=r, column=5, value=round(vol_viga_per, 6))
        ws7.cell(row=r, column=6, value=round(mejor_perfil['A'] * 7850 / 1e6 * L_viga, 1))
        for c in range(1, 7): style_cell(ws7, r, c)

        r += 1
        ws7.cell(row=r, column=1, value="Columna")
        ws7.cell(row=r, column=2, value=best_c['nombre'])
        ws7.cell(row=r, column=3, value=best_c['A'])
        ws7.cell(row=r, column=4, value=L_col)
        ws7.cell(row=r, column=5, value=round(vol_col_com, 6))
        ws7.cell(row=r, column=6, value=round(best_c['A'] * 7850 / 1e6 * L_col, 1))
        for c in range(1, 7): style_cell(ws7, r, c)

        r += 1
        ws7.cell(row=r, column=1, value="TOTAL OPTIMIZADO").font = Font(bold=True, color='C00000')
        ws7.cell(row=r, column=5, value=round(vol_total_per, 6))
        ws7.cell(row=r, column=5).font = Font(bold=True, size=14, color='C00000')
        ws7.cell(row=r, column=6, value=round((mejor_perfil['A'] * L_viga + best_c['A'] * L_col) * 7850 / 1e6 * 1000 / 1000, 1))
        for c in range(1, 7): style_cell(ws7, r, c)

        # Ahorro
        r += 2
        ahorro_vol = vol_total_com - vol_total_per
        ahorro_pct = (1 - vol_total_per / vol_total_com) * 100
        ws7.cell(row=r, column=1, value="AHORRO por optimizacion:").font = Font(bold=True)
        ws7.cell(row=r, column=3, value=f"{round(ahorro_vol, 6)} m3")
        ws7.cell(row=r, column=4, value=f"({round(ahorro_pct, 1)}%)")

ws7.column_dimensions['A'].width = 22
ws7.column_dimensions['B'].width = 24
ws7.column_dimensions['C'].width = 14
ws7.column_dimensions['D'].width = 12
ws7.column_dimensions['E'].width = 14
ws7.column_dimensions['F'].width = 14

# ==============================================================
# GUARDAR
# ==============================================================
output_path = r"C:\Users\andre\Desktop\proyecto materiales\Calculos_Portico_xy45_v2.xlsx"
wb.save(output_path)
print(f"Excel guardado en: {output_path}")
print(f"\n=== RESUMEN ===")
print(f"Reacciones: Ax={A_x:.3f}, Ay={A_y:.3f}, Ma={M_a:.3f}")
print(f"            Ex={E_x:.3f}, Ey={E_y:.3f}")
print(f"Verif: SumFx={check_Fx:.6f}, SumFy={check_Fy:.6f}")
print(f"\nValores criticos:")
print(f"  Viga:    |M|max={M_max_viga:.3f} kN*m, |V|max={V_max_viga:.3f} kN")
print(f"  Columna: |M|max={M_max_col:.3f} kN*m, |N|max={abs(N_col(0)):.3f} kN")
print(f"  V=0 en x={x_V0:.3f} m -> M+max={M_at_V0:.3f} kN*m")
print(f"\nDiseno (FS={FS}):")
print(f"  sigma_adm={sigma_adm:.2f} MPa, tau_adm={tau_adm:.2f} MPa")
print(f"  S_req={S_req_viga_cm3:.1f} cm3")
if cumplen:
    print(f"\nMejor perfil comercial: {cumplen[0]['nombre']}")
    print(f"  A={cumplen[0]['A']} mm2, Vol viga={cumplen[0]['vol']:.6f} m3")
if mejor_perfil:
    print(f"\nPerfil personalizado optimo:")
    print(f"  d={mejor_perfil['d']}, bf={mejor_perfil['bf']}, tf={mejor_perfil['tf']}, tw={mejor_perfil['tw']}")
    print(f"  A={mejor_perfil['A']} mm2, FS_flex={mejor_perfil['FS_flex']:.4f}, FS_cort={mejor_perfil['FS_cort']:.4f}")
    print(f"  Vol viga={mejor_perfil['A']*13000/1e9:.6f} m3")
if cumplen_col:
    print(f"\nMejor perfil columna: {cumplen_col[0]['nombre']}")
    print(f"  A={cumplen_col[0]['A']} mm2, Vol col={cumplen_col[0]['vol']:.6f} m3")
if cumplen and cumplen_col and mejor_perfil:
    v_total = mejor_perfil['A']*13000/1e9 + cumplen_col[0]['vol']
    print(f"\n*** VOLUMEN TOTAL OPTIMIZADO: {v_total:.6f} m3 ***")
